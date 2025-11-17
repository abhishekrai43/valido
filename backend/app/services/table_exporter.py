"""
Table Exporter Service
Exports extracted table data to Excel and PDF formats.
"""
from typing import List, Dict, Any, Optional
import os
from datetime import datetime
from app.utils.logger import get_logger

logger = get_logger("TableExporter")


class TableExporter:
    """Handles exporting table data to various formats."""
    
    @staticmethod
    def export_to_excel(tables: List[Dict[str, Any]], output_path: str, sheet_name: str = "Tables") -> bool:
        """
        Export table data to Excel format.
        
        Args:
            tables: List of table dictionaries from TableExtractor
            output_path: Path to save Excel file
            sheet_name: Name for the Excel sheet
        
        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # If multiple tables, create separate sheets
            if len(tables) > 1:
                for idx, table in enumerate(tables, 1):
                    sheet_name_with_idx = f"Table_{idx}_P{table['page']}"
                    ws = wb.create_sheet(sheet_name_with_idx)
                    TableExporter._write_table_to_sheet(ws, table)
            else:
                # Single table
                ws = wb.create_sheet(sheet_name)
                if tables:
                    TableExporter._write_table_to_sheet(ws, tables[0])
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Exported {len(tables)} table(s) to Excel: {output_path}")
            return True
            
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Failed to export to Excel: {str(e)}")
            return False
    
    @staticmethod
    def _write_table_to_sheet(ws, table: Dict[str, Any]):
        """
        Write table data to Excel worksheet with formatting.
        
        Args:
            ws: openpyxl worksheet
            table: Table dictionary
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Add metadata at top
        ws.append([f"Source: Page {table['page']}, Table {table['table_index']}"])
        ws.append([f"Dimensions: {table['rows']} rows × {table['columns']} columns"])
        ws.append([f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([])  # Empty row
        
        # Style for metadata
        for row in range(1, 4):
            ws[f"A{row}"].font = Font(italic=True, color="666666")
        
        # Write table data
        if table['raw']:
            # Header row
            header_row = []
            for cell in table['headers']:
                header_row.append(str(cell))
            ws.append(header_row)
            
            # Style header
            header_row_num = ws.max_row
            for col_num, cell in enumerate(ws[header_row_num], 1):
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for data_row in table['data']:
                ws.append([str(cell) for cell in data_row])
            
            # Add borders to all table cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            start_row = header_row_num
            end_row = ws.max_row
            for row in range(start_row, end_row + 1):
                for col in range(1, table['columns'] + 1):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Auto-adjust column widths
            for col_num in range(1, table['columns'] + 1):
                max_length = 0
                for row_num in range(start_row, end_row + 1):
                    cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
                    max_length = max(max_length, len(cell_value))
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width
    
    @staticmethod
    def export_to_pdf(tables: List[Dict[str, Any]], output_path: str, title: str = "Extracted Tables") -> bool:
        """
        Export table data to PDF format.
        
        Args:
            tables: List of table dictionaries from TableExtractor
            output_path: Path to save PDF file
            title: Title for the PDF document
        
        Returns:
            True if successful, False otherwise
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, 
                Spacer, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            # Create PDF document
            doc = SimpleDocTemplate(
                output_path,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )
            
            # Container for PDF elements
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#4CAF50'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Add each table
            for idx, table in enumerate(tables, 1):
                # Table header
                table_title = f"Table {table['table_index']} from Page {table['page']}"
                table_header_style = ParagraphStyle(
                    'TableHeader',
                    parent=styles['Heading2'],
                    fontSize=12,
                    textColor=colors.HexColor('#333333'),
                    spaceAfter=6
                )
                elements.append(Paragraph(table_title, table_header_style))
                
                # Metadata
                metadata = f"Dimensions: {table['rows']} rows × {table['columns']} columns"
                metadata_style = ParagraphStyle(
                    'Metadata',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#666666'),
                    spaceAfter=6
                )
                elements.append(Paragraph(metadata, metadata_style))
                
                # Create table data for ReportLab
                if table['raw']:
                    table_data = []
                    
                    # Add all rows (including header)
                    for row in table['raw']:
                        table_data.append([str(cell) for cell in row])
                    
                    # Create ReportLab Table
                    pdf_table = Table(table_data)
                    
                    # Style the table
                    table_style = TableStyle([
                        # Header row styling
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        
                        # Data rows styling
                        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        
                        # Grid
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        
                        # Alternating row colors
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
                    ])
                    
                    pdf_table.setStyle(table_style)
                    elements.append(pdf_table)
                
                # Add spacing between tables
                if idx < len(tables):
                    elements.append(Spacer(1, 0.3*inch))
                    # Page break for new table if more than 2 tables
                    if len(tables) > 2 and idx < len(tables):
                        elements.append(PageBreak())
            
            # Footer
            elements.append(Spacer(1, 0.3*inch))
            footer_text = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} by Valido"
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#999999'),
                alignment=TA_CENTER
            )
            elements.append(Paragraph(footer_text, footer_style))
            
            # Build PDF
            doc.build(elements)
            logger.info(f"Exported {len(tables)} table(s) to PDF: {output_path}")
            return True
            
        except ImportError as ie:
            logger.error(f"reportlab not installed. Install with: pip install reportlab")
            return False
        except Exception as e:
            logger.error(f"Failed to export to PDF: {str(e)}")
            return False
    
    @staticmethod
    def export_to_csv(table: Dict[str, Any], output_path: str) -> bool:
        """
        Export single table to CSV format.
        
        Args:
            table: Table dictionary from TableExtractor
            output_path: Path to save CSV file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            import csv
            
            with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write all rows (including header)
                if table['raw']:
                    for row in table['raw']:
                        writer.writerow(row)
            
            logger.info(f"Exported table to CSV: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export to CSV: {str(e)}")
            return False
