# app/tasks/report_generator.py
"""
Excel and PDF Report Generation Module
Handles creation of professional Excel and PDF reports with formatting.
"""
# pyright: reportOptionalCall=false, reportOptionalMemberAccess=false

import os
from typing import List, Dict, Optional
from datetime import datetime

# Excel dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    openpyxl = None  # type: ignore
    Font = None  # type: ignore
    PatternFill = None  # type: ignore
    Alignment = None  # type: ignore
    Border = None  # type: ignore
    Side = None  # type: ignore
    get_column_letter = None  # type: ignore

# PDF dependencies
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    PDF_REPORT_AVAILABLE = True
except ImportError:
    PDF_REPORT_AVAILABLE = False
    letter = None  # type: ignore
    getSampleStyleSheet = None  # type: ignore
    ParagraphStyle = None  # type: ignore
    inch = None  # type: ignore
    colors = None  # type: ignore
    TA_CENTER = None  # type: ignore
    SimpleDocTemplate = None  # type: ignore
    Table = None  # type: ignore
    TableStyle = None  # type: ignore
    Paragraph = None  # type: ignore
    Spacer = None  # type: ignore


def draw_elegant_background(canvas, doc):
    """Draw elegant light blue background for PDF reports."""
    width, height = doc.pagesize
    canvas.saveState()
    
    # Light blue gradient background
    canvas.setFillColorRGB(0.85, 0.93, 0.98)
    canvas.ellipse(-width * 0.4, height * 0.05, width * 1.2, height * 0.7, fill=True, stroke=False)
    
    canvas.setFillColorRGB(0.70, 0.85, 0.95)
    canvas.ellipse(-width * 0.3, height * 0.2, width * 1.1, height * 0.8, fill=True, stroke=False)
    
    canvas.restoreState()


def draw_valido_footer(canvas, doc):
    """Draw Valido branding footer on each PDF page."""
    width, height = doc.pagesize
    canvas.saveState()
    
    # Footer line
    canvas.setStrokeColorRGB(0.7, 0.7, 0.7)
    canvas.setLineWidth(0.5)
    canvas.line(50, 30, width - 50, 30)
    
    # Valido branding
    canvas.setFont('Helvetica-Bold', 8)
    canvas.setFillColorRGB(0.3, 0.3, 0.3)
    canvas.drawString(50, 18, f"Validated by Valido™")
    
    # Date and page number
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(width - 50, 18, f"{datetime.utcnow().strftime('%B %d, %Y')} | Page {doc.page}")
    
    canvas.restoreState()


def generate_excel_report(csv_rows: List[Dict], output_path: str, timestamp: str) -> Optional[str]:
    """
    Generate Excel report with professional formatting.
    
    Args:
        csv_rows: List of dictionaries containing validation results
        output_path: Directory path where Excel file will be saved
        timestamp: Timestamp string for filename
        
    Returns:
        Excel filename if successful, None otherwise
    """
    if not EXCEL_AVAILABLE or not csv_rows:
        return None
    
    try:
        excel_filename = f'valido_results_{timestamp}.xlsx'
        excel_path = os.path.join(output_path, excel_filename)
        
        wb = openpyxl.Workbook()  # type: ignore
        
        # Remove default sheet and create Summary sheet
        wb.remove(wb.active)  # type: ignore
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Summary sheet styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # type: ignore
        header_font = Font(bold=True, color="FFFFFF", size=14)  # type: ignore
        
        ws_summary['A1'] = "Valido Validation Report"
        ws_summary['A1'].font = Font(bold=True, size=18, color="0066CC")  # type: ignore
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A3'] = "Report Date:"
        ws_summary['B3'] = datetime.utcnow().strftime('%B %d, %Y at %I:%M %p UTC')
        ws_summary['A4'] = "Total Files Processed:"
        ws_summary['B4'] = len(csv_rows)
        
        # Count statuses
        success_count = sum(1 for r in csv_rows if r.get('Status') == 'Success')
        error_count = sum(1 for r in csv_rows if r.get('Status') == 'Error')
        scanned_count = sum(1 for r in csv_rows if r.get('Status') == 'Scanned PDF')
        
        ws_summary['A5'] = "Successful:"
        ws_summary['B5'] = success_count
        ws_summary['A6'] = "Errors:"
        ws_summary['B6'] = error_count
        ws_summary['A7'] = "Scanned PDFs:"
        ws_summary['B7'] = scanned_count
        
        # Count signature types
        signed_count = sum(1 for r in csv_rows if r.get('Signed') == 'Yes')
        digital_signed_count = sum(1 for r in csv_rows if r.get('Digital Signed') == 'Yes')
        
        ws_summary['A9'] = "Documents with Text Signature:"
        ws_summary['B9'] = signed_count
        ws_summary['A10'] = "Documents with Digital Signature:"
        ws_summary['B10'] = digital_signed_count
        
        # Format summary cells
        for row in range(3, 11):
            ws_summary[f'A{row}'].font = Font(bold=True)  # type: ignore
            ws_summary[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # type: ignore
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # Details sheet
        ws_details = wb.create_sheet("Details")
        
        # Write headers
        if csv_rows:
            fieldnames = list(csv_rows[0].keys())
            ws_details.append(fieldnames)
            
            # Style header row
            for cell in ws_details[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # type: ignore
            
            # Write data rows
            for row_data in csv_rows:
                row_values = [row_data.get(field, '') for field in fieldnames]
                ws_details.append(row_values)
            
            # Format data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Apply formatting to all cells
            for row_idx, row in enumerate(ws_details.iter_rows(min_row=2, max_row=len(csv_rows) + 1), start=2):
                row_status = csv_rows[row_idx - 2].get('Status', '')
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Highlight based on status
                    if row_status == 'Error':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif row_status == 'Scanned PDF':
                        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                    elif row_status == 'Success':
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws_details.freeze_panes = 'A2'
        
        wb.save(excel_path)
        return excel_filename
        
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None


def generate_pdf_summary(csv_rows: List[Dict], output_path: str, timestamp: str, rules: Optional[Dict] = None, job_metadata: Optional[Dict] = None) -> Optional[str]:
    """
    Generate professional PDF summary report.
    
    Args:
        csv_rows: List of dictionaries containing validation results
        output_path: Directory path where PDF file will be saved
        timestamp: Timestamp string for filename
        rules: Optional rules dictionary for additional context
        job_metadata: Optional job metadata (name, input/output paths, schedule)
        
    Returns:
        PDF filename if successful, None otherwise
    """
    if not PDF_REPORT_AVAILABLE or not csv_rows:
        return None
    
    try:
        pdf_filename = f'valido_summary_{timestamp}.pdf'
        pdf_path = os.path.join(output_path, pdf_filename)
        
        margin = 50
        doc = SimpleDocTemplate(
            pdf_path, 
            pagesize=letter,
            rightMargin=margin, 
            leftMargin=margin,
            topMargin=margin + 20, 
            bottomMargin=margin + 20
        )
        
        styles = getSampleStyleSheet()
        
        # Light blue color scheme
        title_color = colors.HexColor('#5DADE2')
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=28,
            textColor=title_color,
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=title_color,
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            textColor=colors.black
        )
        
        flowables = []
        
        # Title
        flowables.append(Paragraph("Valido Validation Report", title_style))
        flowables.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%B %d, %Y at %I:%M %p UTC')}", 
            subtitle_style
        ))
        
        # Job Information (if available)
        if job_metadata:
            flowables.append(Paragraph("Job Information", heading_style))
            
            job_info_data = []
            
            if job_metadata.get('name'):
                job_info_data.append(['Job Name', job_metadata['name']])
            
            if job_metadata.get('input_path'):
                job_info_data.append(['Input Folder', job_metadata['input_path']])
            
            if job_metadata.get('output_path'):
                job_info_data.append(['Output Folder', job_metadata['output_path']])
            
            if job_metadata.get('execution_type'):
                job_info_data.append(['Execution Type', job_metadata['execution_type']])
            
            if job_metadata.get('schedule_times'):
                job_info_data.append(['Schedule', job_metadata['schedule_times']])
            
            if job_metadata.get('ruleset_name'):
                job_info_data.append(['Ruleset', job_metadata['ruleset_name']])
            
            if job_info_data:
                job_info_table = Table(job_info_data, colWidths=[2*inch, 4*inch])
                job_info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.95, 1.0)),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 10),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ]))
                
                flowables.append(job_info_table)
                flowables.append(Spacer(1, 20))
        
        # Executive Summary
        flowables.append(Paragraph("Executive Summary", heading_style))
        
        total_files = len(csv_rows)
        success_count = sum(1 for r in csv_rows if r.get('Status') == 'Success')
        error_count = sum(1 for r in csv_rows if r.get('Status') == 'Error')
        scanned_count = sum(1 for r in csv_rows if r.get('Status') == 'Scanned PDF')
        success_rate = (success_count / total_files * 100) if total_files > 0 else 0
        
        summary_data = [
            ['Metric', 'Count', 'Percentage'],
            ['Total Files Processed', str(total_files), '100%'],
            ['Successfully Validated', str(success_count), f'{success_rate:.1f}%'],
            ['Errors', str(error_count), f'{(error_count/total_files*100) if total_files > 0 else 0:.1f}%'],
            ['Scanned PDFs (Not Supported)', str(scanned_count), f'{(scanned_count/total_files*100) if total_files > 0 else 0:.1f}%'],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), title_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.95, 0.97, 1.0), colors.lightgrey]),
        ]))
        
        flowables.append(summary_table)
        flowables.append(Spacer(1, 20))
        
        # Validation Details
        signed_count = sum(1 for r in csv_rows if r.get('Signed') == 'Yes')
        digital_signed_count = sum(1 for r in csv_rows if r.get('Digital Signed') == 'Yes')
        dated_count = sum(1 for r in csv_rows if r.get('Dated') == 'Yes')
        
        flowables.append(Paragraph("Validation Details", heading_style))
        
        validation_data = [
            ['Validation Check', 'Results'],
            ['Documents with Text Signature', f'{signed_count} of {total_files}'],
            ['Documents with Digital Signature', f'{digital_signed_count} of {total_files}'],
            ['Documents with Date', f'{dated_count} of {total_files}'],
        ]
        
        validation_table = Table(validation_data, colWidths=[3.5*inch, 2.5*inch])
        validation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), title_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        flowables.append(validation_table)
        flowables.append(Spacer(1, 20))
        
        # Detailed file-by-file results (limited to first 20 files)
        if total_files <= 20:
            flowables.append(Paragraph("Detailed Results by File", heading_style))
            
            for idx, row in enumerate(csv_rows, 1):
                filename = row.get('Filename', 'Unknown')
                status = row.get('Status', 'Unknown')
                
                # File header
                file_header = f"{idx}. {filename}"
                flowables.append(Paragraph(file_header, ParagraphStyle(
                    'FileHeader',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=colors.HexColor('#2874A6'),
                    spaceAfter=8,
                    spaceBefore=12,
                    fontName='Helvetica-Bold'
                )))
                
                # Build details list
                details = []
                details.append(f"<b>Status:</b> {status}")
                
                if row.get('Signed'):
                    details.append(f"<b>Text Signature:</b> {row.get('Signed')}")
                    if row.get('Signature Details'):
                        sig_details = row.get('Signature Details', '')
                        if sig_details:
                            details.append(f"  → {sig_details[:80]}")
                
                if row.get('Digital Signed'):
                    details.append(f"<b>Digital Signature:</b> {row.get('Digital Signed')}")
                
                if row.get('Dated'):
                    details.append(f"<b>Date Found:</b> {row.get('Dated')}")
                    if row.get('Date Found'):
                        details.append(f"  → {row.get('Date Found')}")
                
                # Add extracted fields (exclude internal fields)
                excluded_keys = [
                    'Filename', 'Status', 'Signed', 'Digital Signed', 'Dated', 
                    'Signature Details', 'Date Found', 'Error Details', 'Signature Type', 
                    'Digital Signature Details', 'Actual Page Count', 'validation_report'
                ]
                for key, value in row.items():
                    if key not in excluded_keys:
                        if value and value != 'N/A' and value != '':
                            details.append(f"<b>{key}:</b> {value}")
                
                if row.get('Error Details'):
                    err_details = row.get('Error Details', '')
                    if err_details:
                        details.append(f"<b>Error:</b> {err_details[:100]}")
                
                # Add details as paragraph
                details_text = '<br/>'.join(details)
                flowables.append(Paragraph(details_text, ParagraphStyle(
                    'FileDetails',
                    parent=styles['Normal'],
                    fontSize=10,
                    leftIndent=20,
                    spaceAfter=10,
                    leading=14
                )))
            
            flowables.append(Spacer(1, 10))
        
        # Failed/Error files (if any)
        if error_count > 0 or scanned_count > 0:
            flowables.append(Paragraph("Files Requiring Attention", heading_style))
            
            problem_files = []
            for r in csv_rows:
                if r.get('Status') in ['Error', 'Scanned PDF']:
                    problem_files.append([
                        r.get('Filename', 'Unknown'),
                        r.get('Status', 'Unknown'),
                        r.get('Error Details', 'N/A')[:50]  # Truncate long errors
                    ])
            
            if problem_files:
                problem_data = [['Filename', 'Status', 'Details']] + problem_files[:20]  # Limit to first 20
                
                problem_table = Table(problem_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
                problem_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC143C')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                flowables.append(problem_table)
                
                if len(problem_files) > 20:
                    flowables.append(Spacer(1, 10))
                    flowables.append(Paragraph(
                        f"<i>Note: Showing first 20 of {len(problem_files)} files requiring attention. See Excel file for complete details.</i>",
                        normal_style
                    ))
        
        # Footer note
        flowables.append(Spacer(1, 30))
        flowables.append(Paragraph(
            "<i>This report was automatically generated by Valido. For detailed results, please refer to the Excel file included in this package.</i>",
            normal_style
        ))
        
        # Build PDF with background and footer
        doc.build(
            flowables,
            onFirstPage=lambda c, d: (draw_elegant_background(c, d), draw_valido_footer(c, d)),
            onLaterPages=lambda c, d: (draw_elegant_background(c, d), draw_valido_footer(c, d))
        )
        
        return pdf_filename
        
    except Exception as e:
        print(f"Error generating PDF summary: {e}")
        import traceback
        traceback.print_exc()
        return None
