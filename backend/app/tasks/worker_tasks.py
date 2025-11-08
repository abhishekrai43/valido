# app/tasks/worker_tasks.py
import time
import io
import zipfile
import os
import csv
import json
import re
from typing import List, Dict, Any, Optional, Tuple, Callable
from datetime import datetime

from app.services.parser import extract_text_from_bytes, is_valid_pdf
from app.services.validator import validate_text
from PyPDF2 import PdfReader

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    PDF_REPORT_AVAILABLE = True
except ImportError:
    PDF_REPORT_AVAILABLE = False


# Professional color palettes for PDF reports
BACKGROUND_COLORS = [
    [(0.85, 0.92, 0.98), (0.7, 0.85, 0.95)],  # Blue tones
    [(0.90, 0.95, 0.90), (0.75, 0.88, 0.75)],  # Green tones
    [(0.95, 0.88, 0.90), (0.88, 0.75, 0.80)],  # Rose tones
    [(0.95, 0.92, 0.85), (0.88, 0.82, 0.70)],  # Warm tones
    [(0.92, 0.90, 0.95), (0.80, 0.75, 0.88)],  # Purple tones
]

HEADER_COLORS = [
    colors.HexColor('#0066CC'),  # Professional Blue
    colors.HexColor('#2D5F2E'),  # Forest Green
    colors.HexColor('#8B4513'),  # Saddle Brown
    colors.HexColor('#4B0082'),  # Indigo
    colors.HexColor('#DC143C'),  # Crimson
]


def draw_elegant_background(canvas, doc):
    """Draw random elegant oval patterns as background."""
    width, height = doc.pagesize
    canvas.saveState()
    
    # Choose random color palette
    import random
    color_palette = random.choice(BACKGROUND_COLORS)
    
    # Draw two overlapping ovals with lighter shades
    canvas.setFillColorRGB(*color_palette[0])
    canvas.ellipse(-width * 0.4, height * 0.05, width * 1.2, height * 0.7, fill=True, stroke=False)
    
    canvas.setFillColorRGB(*color_palette[1])
    canvas.ellipse(-width * 0.3, height * 0.2, width * 1.1, height * 0.8, fill=True, stroke=False)
    
    canvas.restoreState()


def draw_valido_footer(canvas, doc):
    """Draw Valido branding footer on each page."""
    width, height = doc.pagesize
    canvas.saveState()
    
    # Footer line
    canvas.setStrokeColorRGB(0.7, 0.7, 0.7)
    canvas.setLineWidth(0.5)
    canvas.line(50, 30, width - 50, 30)
    
    # Valido branding
    canvas.setFont('Helvetica-Bold', 8)
    canvas.setFillColorRGB(0.3, 0.3, 0.3)
    canvas.drawString(50, 18, f"Validated by Valido™")
    
    # Date and page number
    canvas.setFont('Helvetica', 8)
    canvas.drawRightString(width - 50, 18, f"{datetime.utcnow().strftime('%B %d, %Y')} | Page {doc.page}")
    
    canvas.restoreState()


def generate_excel_report(csv_rows: List[Dict], output_path: str, timestamp: str) -> Optional[str]:
    """Generate Excel report with professional formatting."""
    if not EXCEL_AVAILABLE or not csv_rows:
        return None
    
    try:
        excel_filename = f'valido_results_{timestamp}.xlsx'
        excel_path = os.path.join(output_path, excel_filename)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet and create Summary sheet
        wb.remove(wb.active)
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Summary sheet styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        ws_summary['A1'] = "Valido Validation Report"
        ws_summary['A1'].font = Font(bold=True, size=18, color="0066CC")
        ws_summary.merge_cells('A1:B1')
        
        ws_summary['A3'] = "Report Date:"
        ws_summary['B3'] = datetime.utcnow().strftime('%B %d, %Y at %I:%M %p UTC')
        ws_summary['A4'] = "Total Files Processed:"
        ws_summary['B4'] = len(csv_rows)
        
        # Count statuses
        success_count = sum(1 for r in csv_rows if r.get('Status') == 'Success')
        error_count = sum(1 for r in csv_rows if r.get('Status') == 'Error')
        scanned_count = sum(1 for r in csv_rows if r.get('Status') == 'Scanned PDF')
        
        ws_summary['A5'] = "Successful:"
        ws_summary['B5'] = success_count
        ws_summary['A6'] = "Errors:"
        ws_summary['B6'] = error_count
        ws_summary['A7'] = "Scanned PDFs:"
        ws_summary['B7'] = scanned_count
        
        # Count signature types
        signed_count = sum(1 for r in csv_rows if r.get('Signed') == 'Yes')
        digital_signed_count = sum(1 for r in csv_rows if r.get('Digital Signed') == 'Yes')
        
        ws_summary['A9'] = "Documents with Text Signature:"
        ws_summary['B9'] = signed_count
        ws_summary['A10'] = "Documents with Digital Signature:"
        ws_summary['B10'] = digital_signed_count
        
        # Format summary cells
        for row in range(3, 11):
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # Details sheet
        ws_details = wb.create_sheet("Details")
        
        # Write headers
        if csv_rows:
            fieldnames = list(csv_rows[0].keys())
            ws_details.append(fieldnames)
            
            # Style header row
            for cell in ws_details[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data rows
            for row_data in csv_rows:
                row_values = [row_data.get(field, '') for field in fieldnames]
                ws_details.append(row_values)
            
            # Format data cells
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Apply formatting to all cells
            for row_idx, row in enumerate(ws_details.iter_rows(min_row=2, max_row=len(csv_rows) + 1), start=2):
                row_status = csv_rows[row_idx - 2].get('Status', '')
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Highlight based on status
                    if row_status == 'Error':
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif row_status == 'Scanned PDF':
                        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                    elif row_status == 'Success':
                        cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws_details.freeze_panes = 'A2'
        
        wb.save(excel_path)
        return excel_filename
        
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        return None


def generate_pdf_summary(csv_rows: List[Dict], output_path: str, timestamp: str, rules: Optional[Dict] = None) -> Optional[str]:
    """Generate professional PDF summary with random elegant design."""
    if not PDF_REPORT_AVAILABLE or not csv_rows:
        return None
    
    try:
        import random
        
        pdf_filename = f'valido_summary_{timestamp}.pdf'
        pdf_path = os.path.join(output_path, pdf_filename)
        
        margin = 50
        doc = SimpleDocTemplate(
            pdf_path, 
            pagesize=letter,
            rightMargin=margin, 
            leftMargin=margin,
            topMargin=margin + 20, 
            bottomMargin=margin + 20
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles with random colors
        title_color = random.choice(HEADER_COLORS)
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=28,
            textColor=title_color,
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#666666'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=title_color,
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            textColor=colors.black
        )
        
        flowables = []
        
        # Title
        flowables.append(Paragraph("Valido Validation Report", title_style))
        flowables.append(Paragraph(
            f"Generated on {datetime.utcnow().strftime('%B %d, %Y at %I:%M %p UTC')}", 
            subtitle_style
        ))
        
        # Executive Summary
        flowables.append(Paragraph("Executive Summary", heading_style))
        
        total_files = len(csv_rows)
        success_count = sum(1 for r in csv_rows if r.get('Status') == 'Success')
        error_count = sum(1 for r in csv_rows if r.get('Status') == 'Error')
        scanned_count = sum(1 for r in csv_rows if r.get('Status') == 'Scanned PDF')
        success_rate = (success_count / total_files * 100) if total_files > 0 else 0
        
        summary_data = [
            ['Metric', 'Count', 'Percentage'],
            ['Total Files Processed', str(total_files), '100%'],
            ['Successfully Validated', str(success_count), f'{success_rate:.1f}%'],
            ['Errors', str(error_count), f'{(error_count/total_files*100) if total_files > 0 else 0:.1f}%'],
            ['Scanned PDFs (Not Supported)', str(scanned_count), f'{(scanned_count/total_files*100) if total_files > 0 else 0:.1f}%'],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), title_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.95, 0.97, 1.0), colors.lightgrey]),
        ]))
        
        flowables.append(summary_table)
        flowables.append(Spacer(1, 20))
        
        # Validation Details
        signed_count = sum(1 for r in csv_rows if r.get('Signed') == 'Yes')
        digital_signed_count = sum(1 for r in csv_rows if r.get('Digital Signed') == 'Yes')
        dated_count = sum(1 for r in csv_rows if r.get('Dated') == 'Yes')
        
        flowables.append(Paragraph("Validation Details", heading_style))
        
        validation_data = [
            ['Validation Check', 'Results'],
            ['Documents with Text Signature', f'{signed_count} of {total_files}'],
            ['Documents with Digital Signature', f'{digital_signed_count} of {total_files}'],
            ['Documents with Date', f'{dated_count} of {total_files}'],
        ]
        
        validation_table = Table(validation_data, colWidths=[3.5*inch, 2.5*inch])
        validation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), title_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        flowables.append(validation_table)
        flowables.append(Spacer(1, 20))
        
        # Failed/Error files (if any)
        if error_count > 0 or scanned_count > 0:
            flowables.append(Paragraph("Files Requiring Attention", heading_style))
            
            problem_files = []
            for r in csv_rows:
                if r.get('Status') in ['Error', 'Scanned PDF']:
                    problem_files.append([
                        r.get('Filename', 'Unknown'),
                        r.get('Status', 'Unknown'),
                        r.get('Error Details', 'N/A')[:50]  # Truncate long errors
                    ])
            
            if problem_files:
                problem_data = [['Filename', 'Status', 'Details']] + problem_files[:20]  # Limit to first 20
                
                problem_table = Table(problem_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
                problem_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC143C')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                flowables.append(problem_table)
                
                if len(problem_files) > 20:
                    flowables.append(Spacer(1, 10))
                    flowables.append(Paragraph(
                        f"<i>Note: Showing first 20 of {len(problem_files)} files requiring attention. See Excel file for complete details.</i>",
                        normal_style
                    ))
        
        # Footer note
        flowables.append(Spacer(1, 30))
        flowables.append(Paragraph(
            "<i>This report was automatically generated by Valido. For detailed results, please refer to the Excel file included in this package.</i>",
            normal_style
        ))
        
        # Build PDF with background and footer
        doc.build(
            flowables,
            onFirstPage=lambda c, d: (draw_elegant_background(c, d), draw_valido_footer(c, d)),
            onLaterPages=lambda c, d: (draw_elegant_background(c, d), draw_valido_footer(c, d))
        )
        
        return pdf_filename
        
    except Exception as e:
        print(f"Error generating PDF summary: {e}")
        import traceback
        traceback.print_exc()
        return None


def process_in_chunks_worker(items: List[Any], chunk_size: int = 50):
    total = len(items or [])
    processed = 0
    outputs = []
    for i in range(0, total, chunk_size):
        chunk = items[i : i + chunk_size]
        time.sleep(0.01)
        chunk_out = [len(x) if isinstance(x, (str, bytes)) else 1 for x in chunk]
        outputs.extend(chunk_out)
        processed += len(chunk)
    return {"status": "completed", "total": total, "processed": processed, "sample": outputs[:10]}


def process_pdfs_sync(
    files: List[Dict],
    rules: Optional[Dict] = None,
    username: Optional[str] = None,
    results_dir: Optional[str] = None,
    progress_callback: Optional[Callable[[str, Dict[str, Any]], None]] = None,
) -> Dict:
    def _emit_progress(state: str, meta: Dict[str, Any]):
        try:
            if progress_callback:
                progress_callback(state, meta)
        except Exception:
            pass

    reports = []
    total_files = 0

    expanded_files: List[Dict[str, Any]] = []
    for entry in files:
        filename = entry.get("filename")
        content = entry.get("content") or b''
        total_files += 1

        is_zip = False
        if filename and filename.lower().endswith(".zip"):
            is_zip = True
        else:
            if isinstance(content, (bytes, bytearray)) and content[:4] == b"PK\x03\x04":
                is_zip = True

        if is_zip:
            try:
                with io.BytesIO(content) as bio:
                    with zipfile.ZipFile(bio) as zf:
                        for zi in zf.infolist():
                            if zi.filename.lower().endswith(".pdf"):
                                with zf.open(zi) as f:
                                    file_bytes = f.read()
                                    expanded_files.append({"filename": zi.filename, "content": file_bytes})
            except Exception:
                reports.append({"filename": filename, "error": "failed to extract zip"})
        else:
            expanded_files.append({"filename": filename, "content": content or b''})

    total = len(expanded_files)
    processed = 0

    _emit_progress("PROGRESS", {"processed": 0, "total": total, "current_file": "", "percent": 0})

    for f in expanded_files:
        fname = f.get("filename")
        content = f.get("content") or b''

        _emit_progress(
            "PROGRESS",
            {
                "processed": processed,
                "total": total,
                "current_file": fname,
                "percent": int((processed / total * 100)) if total > 0 else 0,
            },
        )

        print(f"Processing {fname}: received {len(content)} bytes")
        if content:
            print(f"First 20 bytes: {content[:20]}")

        valid = False
        try:
            valid = is_valid_pdf(content or b'')
            if not valid:
                print(f"PDF validation failed for {fname}: is_valid_pdf returned False")
        except Exception as e:
            print(f"PDF validation exception for {fname}: {type(e).__name__}: {e}")
            valid = False

        if not valid:
            reports.append({"filename": fname, "error": "invalid or corrupted pdf"})
            processed += 1
            _emit_progress(
                "PROGRESS",
                {
                    "processed": processed,
                    "total": total,
                    "current_file": fname,
                    "percent": int((processed / total * 100)) if total > 0 else 0,
                },
            )
            continue

        try:
            text = extract_text_from_bytes(content or b'')
            report = validate_text(text, rules)
            reports.append({"filename": fname, "report": report})
        except Exception as exc:
            reports.append({"filename": fname, "error": f"processing error: {exc}"})

        processed += 1
        _emit_progress(
            "PROGRESS",
            {
                "processed": processed,
                "total": total,
                "current_file": fname,
                "percent": int((processed / total * 100)) if total > 0 else 0,
            },
        )

    task_id = None
    if results_dir:
        results_dir = os.path.abspath(results_dir)
        task_id = os.path.basename(results_dir.rstrip("/\\"))
    else:
        task_id = datetime.utcnow().strftime('%s')
        results_dir = os.path.abspath(os.path.join(os.getcwd(), 'results', task_id))

    os.makedirs(results_dir, exist_ok=True)

    def detect_signed(t: str) -> Tuple[str, str]:
        if not t:
            return 'No', ''
        kws = ['signature', 'signed by', '/s/', 'signatory', 'electronically signed', 'signed:', '/sig/']
        tl = t.lower()
        for k in kws:
            if k in tl:
                idx = tl.find(k)
                snippet = t[max(0, idx-10):min(len(t), idx+50)].strip()
                return 'Yes', snippet[:100]
        return 'No', ''

    def detect_digital_signature(pdf_bytes: bytes) -> Tuple[str, str]:
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            # Check for AcroForm signature fields
            root = reader.trailer.get('/Root', {})
            acroform = root.get('/AcroForm', {}) if isinstance(root, dict) else {}
            fields = acroform.get('/Fields', []) if isinstance(acroform, dict) else []
            for field in fields:
                try:
                    field_obj = field.get_object()
                    if field_obj.get('/FT') == '/Sig':
                        return 'Yes', 'Digital signature field found'
                except Exception:
                    continue
            # Check for /Sig annotation in pages
            for page in reader.pages:
                annots = page.get('/Annots', [])
                for annot_ref in annots:
                    try:
                        annot = annot_ref.get_object()
                        if annot.get('/Subtype') == '/Widget' and annot.get('/FT') == '/Sig':
                            return 'Yes', 'Digital signature annotation found'
                    except Exception:
                        continue
            return 'No', ''
        except Exception as e:
            return 'Error', str(e)

    date_regexes = [
        r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b",
        r"\b\d{4}-\d{1,2}-\d{1,2}\b",
        r"\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]* \d{1,2},? \d{2,4}\b",
    ]
    date_re = re.compile('|'.join(date_regexes), re.IGNORECASE)

    def detect_dated(t: str) -> Tuple[str, str]:
        if not t:
            return 'No', ''
        m = date_re.search(t)
        if m:
            return 'Yes', m.group(0)
        return 'No', ''

    def check_must_contain(t: str, rule: dict) -> Tuple[str, str]:
        if not t or not rule:
            return 'No', ''
        search_text = rule.get('text', '')
        if not search_text:
            return 'No', ''
        case_sensitive = rule.get('case_sensitive', False)
        if case_sensitive:
            found = search_text in t
            if found:
                idx = t.find(search_text)
                snippet = t[max(0, idx-20):min(len(t), idx+len(search_text)+20)].strip()
                return 'Yes', snippet[:100]
        else:
            found = search_text.lower() in t.lower()
            if found:
                idx = t.lower().find(search_text.lower())
                snippet = t[max(0, idx-20):min(len(t), idx+len(search_text)+20)].strip()
                return 'Yes', snippet[:100]
        return 'No', ''

    def check_must_not_contain(t: str, rule: dict) -> Tuple[str, str]:
        if not t or not rule:
            return 'Pass', ''
        search_text = rule.get('text', '')
        if not search_text:
            return 'Pass', ''
        case_sensitive = rule.get('case_sensitive', False)
        if case_sensitive:
            found = search_text in t
            if found:
                idx = t.find(search_text)
                snippet = t[max(0, idx-20):min(len(t), idx+len(search_text)+20)].strip()
                return 'Fail', snippet[:100]
        else:
            found = search_text.lower() in t.lower()
            if found:
                idx = t.lower().find(search_text.lower())
                snippet = t[max(0, idx-20):min(len(t), idx+len(search_text)+20)].strip()
                return 'Fail', snippet[:100]
        return 'Pass', ''

    def check_page_count(content_bytes: bytes, rule: dict) -> Tuple[str, str]:
        if not content_bytes or not rule:
            return 'Unknown', '0'
        try:
            import fitz
            doc = fitz.open(stream=content_bytes, filetype="pdf")
            page_count = len(doc)
            doc.close()
            operator = rule.get('operator', '==')
            value = int(rule.get('value', 0))
            result = False
            if operator == '==':
                result = page_count == value
            elif operator == '>=':
                result = page_count >= value
            elif operator == '<=':
                result = page_count <= value
            return ('Pass' if result else 'Fail'), str(page_count)
        except Exception as e:
            return 'Error', f'Could not count pages: {str(e)[:50]}'

    def extract_field_from_text(text: str, field_name: str, strategy: str = 'first') -> str:
        if not text:
            return ''
        field_lower = field_name.lower().replace('_', ' ')
        patterns = [
            (rf"\b{re.escape(field_lower)}\s*[:]\s*([^\n\r]+)", re.IGNORECASE),
            (rf"\b{re.escape(field_lower)}\s+([A-Za-z0-9₹$€£¥₨,.]+[^\n\r]*?)(?:\s{{2,}}|\n|$)", re.IGNORECASE),
            (rf"\b{re.escape(field_lower)}\s*\n\s*([^\n\r]+)", re.IGNORECASE),
        ]
        all_matches = []
        for pattern, flags in patterns:
            matches = re.finditer(pattern, text, flags)
            for m in matches:
                value = m.group(1).strip()
                if value.isupper() and len(value.split()) <= 3:
                    continue
                value = re.sub(r'\s{2,}', ' ', value)
                value = value.split('\t')[0]
                currency_map = {'₹': 'Rs.', '₨': 'Rs.', '€': 'EUR', '£': 'GBP', '¥': 'JPY', '$': 'USD'}
                for symbol, replacement in currency_map.items():
                    if symbol in value:
                        if value.startswith(symbol):
                            value = replacement + ' ' + value[1:].strip()
                        else:
                            value = value.replace(symbol, replacement)
                if len(value) > 100:
                    value = value[:100].rsplit(' ', 1)[0]
                cleaned_value = value.strip()
                if cleaned_value and cleaned_value not in all_matches:
                    all_matches.append(cleaned_value)
        if not all_matches:
            return ''
        if strategy == 'last':
            return all_matches[-1]
        elif strategy == 'all':
            return ' | '.join(all_matches)
        else:
            return all_matches[0]

    validation_checks = []
    extraction_fields = []
    must_contain_rules = []
    must_not_contain_rules = []
    page_count_rules = None

    if isinstance(rules, dict):
        validations = rules.get('validations', {})
        if isinstance(validations, dict):
            if validations.get('signed'):
                validation_checks.append('signed')
            if validations.get('dated'):
                validation_checks.append('dated')
            if validations.get('signed_and_dated'):
                if 'signed' not in validation_checks:
                    validation_checks.append('signed')
                if 'dated' not in validation_checks:
                    validation_checks.append('dated')
            if validations.get('must_contain'):
                must_contain_rules.append(validations['must_contain'])
            if validations.get('must_not_contain'):
                must_not_contain_rules.append(validations['must_not_contain'])
            if validations.get('page_count'):
                page_count_rules = validations['page_count']
        if rules.get('validate_signed'):
            validation_checks.append('signed')
        if rules.get('validate_dated'):
            validation_checks.append('dated')
        if rules.get('validate_signed_and_dated'):
            if 'signed' not in validation_checks:
                validation_checks.append('signed')
            if 'dated' not in validation_checks:
                validation_checks.append('dated')
        fields = rules.get('fields') or []
        if isinstance(fields, list):
            for field in fields:
                if isinstance(field, dict):
                    extraction_fields.append({'name': field.get('name', ''), 'strategy': field.get('strategy', 'first')})
                elif isinstance(field, str):
                    extraction_fields.append({'name': field, 'strategy': 'first'})

    csv_rows = []
    for entry in reports:
        fname = entry.get('filename')
        err = entry.get('error', '')
        text = ''
        content_bytes = b''
        try:
            for ef in expanded_files:
                if ef.get('filename') == fname:
                    content_bytes = ef.get('content') or b''
                    text = extract_text_from_bytes(content_bytes) or ''
                    break
        except Exception:
            text = ''

        is_scanned = text.startswith('[SCANNED_PDF]')
        scanned_message = ''
        if is_scanned:
            scanned_message = text.replace('[SCANNED_PDF]', '').strip()
            text = ''

        row = {'Filename': fname, 'Status': 'Scanned PDF' if is_scanned else ('Error' if err else 'Success')}
        if is_scanned:
            row['Error Details'] = scanned_message

        # Signature detection (text and digital)
        is_signed, signed_value = detect_signed(text)
        is_digital, digital_value = detect_digital_signature(content_bytes)
        sig_type = []
        if is_signed == 'Yes':
            sig_type.append('Text')
        if is_digital == 'Yes':
            sig_type.append('Digital')
        if not sig_type:
            sig_type_str = 'None'
        else:
            sig_type_str = ', '.join(sig_type)
        row['Signature Type'] = sig_type_str
        row['Signed'] = is_signed
        if signed_value:
            row['Signature Details'] = signed_value
        row['Digital Signed'] = is_digital
        if digital_value:
            row['Digital Signature Details'] = digital_value

        if 'dated' in validation_checks:
            is_dated, date_value = detect_dated(text)
            row['Dated'] = is_dated
            if date_value:
                row['Date Found'] = date_value

        for idx, rule in enumerate(must_contain_rules):
            search_term = rule.get('text', '')
            case_info = ' (case-sensitive)' if rule.get('case_sensitive') else ''
            col_name = f'Contains \"{search_term}\"{case_info}'
            result, snippet = check_must_contain(text, rule)
            row[col_name] = result
            if snippet:
                row[f'{col_name} - Found'] = snippet

        for idx, rule in enumerate(must_not_contain_rules):
            search_term = rule.get('text', '')
            case_info = ' (case-sensitive)' if rule.get('case_sensitive') else ''
            col_name = f'Does NOT Contain \"{search_term}\"{case_info}'
            result, snippet = check_must_not_contain(text, rule)
            row[col_name] = result
            if snippet and result == 'Fail':
                row[f'{col_name} - Found'] = snippet

        if page_count_rules:
            operator = page_count_rules.get('operator', '==')
            value = page_count_rules.get('value', 0)
            op_text = {'==': 'exactly', '>=': 'at least', '<=': 'at most'}.get(operator, operator)
            col_name = f'Page Count ({op_text} {value})'
            result, actual_count = check_page_count(content_bytes, page_count_rules)
            row[col_name] = result
            row['Actual Page Count'] = actual_count

        for field in extraction_fields:
            field_name = field.get('name', field) if isinstance(field, dict) else field
            field_strategy = field.get('strategy', 'first') if isinstance(field, dict) else 'first'
            field_display = field_name.replace('_', ' ').title()
            extracted_value = extract_field_from_text(text, field_name, field_strategy) if text else ''
            row[field_display] = extracted_value

        if err:
            row['Error Details'] = err

        csv_rows.append(row)

    if csv_rows:
        fieldnames = list(csv_rows[0].keys())
    else:
        fieldnames = ['Filename', 'Status']

    timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    csv_filename = f'valido_results_{timestamp}.csv'
    csv_path = os.path.join(results_dir, csv_filename)
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as cf:
        writer = csv.DictWriter(cf, fieldnames=fieldnames)
        writer.writeheader()
        for row in csv_rows:
            safe_row = {}
            for key, value in row.items():
                if value is None:
                    safe_row[key] = ''
                elif isinstance(value, bytes):
                    safe_row[key] = value.decode('utf-8', errors='replace')
                else:
                    safe_row[key] = str(value)
            writer.writerow(safe_row)

    # Generate Excel report
    excel_filename = generate_excel_report(csv_rows, results_dir, timestamp)
    
    # Generate PDF summary
    pdf_filename = generate_pdf_summary(csv_rows, results_dir, timestamp, rules)
    
    # Generate JSON for advanced users
    json_filename = f'valido_results_{timestamp}.json'
    json_path = os.path.join(results_dir, json_filename)
    with open(json_path, 'w', encoding='utf-8') as jf:
        json.dump({
            'generated_at': datetime.utcnow().isoformat(),
            'total_files': total,
            'processed': processed,
            'results': csv_rows
        }, jf, indent=2, ensure_ascii=False)

    if username:
        try:
            from app.db import get_session
            from app.models import User
            from sqlmodel import select
            with get_session() as session:
                u = session.exec(select(User).where(User.username == username)).first()
                if not u:
                    u = User(username=username, total_processed=processed)
                    session.add(u)
                else:
                    u.total_processed = (u.total_processed or 0) + processed
                    session.add(u)
                session.commit()
        except Exception:
            pass

    return {
        'status': 'completed',
        'total': total,
        'processed': processed,
        'files': reports,
        'result_files': {
            'csv': f'/api/v1/tasks/{task_id}/result.csv',
            'zip': f'/api/v1/tasks/{task_id}/results.zip'
        }
    }


# backward-compatible export name
process_pdfs = process_pdfs_sync
