"""Tests for Excel report skipped-files sheet.

These tests are intentionally lightweight and don't require real PDFs.
They validate that failed/scanned rows are surfaced to the user in a
human-friendly way in the generated Excel workbook.
"""

import os

from app.tasks.report_generator import generate_excel_report


def test_excel_contains_skipped_files_sheet(tmp_path):
    rows = [
        {
            "Filename": "good.pdf",
            "Status": "Success",
            "Signed": "No",
            "Digital Signed": "No",
        },
        {
            "Filename": "scanned.pdf",
            "Status": "Scanned PDF",
            "Error Details": "[SCANNED_PDF] This PDF appears to be a scan",
        },
        {
            "Filename": "corrupt.pdf",
            "Status": "Error",
            "Error Details": "invalid or corrupted pdf",
        },
    ]

    out_dir = str(tmp_path)
    filename = generate_excel_report(rows, out_dir, timestamp="20260119_000000")
    assert filename

    excel_path = os.path.join(out_dir, filename)
    assert os.path.exists(excel_path)

    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    assert "Skipped Files" in wb.sheetnames

    ws = wb["Skipped Files"]
    # header row
    assert ws["A1"].value == "Filename"
    assert ws["B1"].value == "Reason"

    # confirm the two skipped rows show up
    values = [ws[f"A{i}"].value for i in range(2, ws.max_row + 1)]
    assert "scanned.pdf" in values
    assert "corrupt.pdf" in values
